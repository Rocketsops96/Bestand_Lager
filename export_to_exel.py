import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import filedialog, Tk  # добавлен импорт для диалогового окна
import regbase

def export_to_excel():

    # Подключаемся к базе данных BAU
    conn = regbase.create_conn()
    cursor = conn.cursor()

    # Выполняем запрос к базе данных и получаем данные
    cursor.execute("SELECT * FROM lager_bestand")
    data = cursor.fetchall()

    # Создаем новую книгу Excel и выбираем активный лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Стили для заголовков
    header_style = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Записываем заголовки столбцов в Excel
    columns = [description[0] for description in cursor.description]
    for col_num, col_title in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=col_title)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_alignment

    # Стили для данных
    data_style = Font(name='Calibri')
    data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='left', vertical='center')

    # Записываем данные из базы данных в Excel
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_data)
            cell.font = data_style
            cell.border = data_border
            cell.alignment = data_alignment

    # Открываем диалоговое окно для выбора места сохранения файла
    root = Tk()
    root.withdraw()  # Скрываем основное окно

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    # Сохраняем книгу Excel в выбранное место
    if file_path:
        workbook.save(file_path)
        

    # Закрываем соединение с базой данных
    conn.close()


